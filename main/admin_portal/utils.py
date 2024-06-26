from . import models
import pandas as pd
from .models import Students
import os 
from django.conf import settings 
from openpyxl import load_workbook

def reconciliation(): 
    id  = "reconciliation.xlsx"
    excel_file_path = os.path.join(settings.BASE_DIR, "static", "excel", id)
    wb = load_workbook(excel_file_path)
    ws = wb.active 
    for index,student in enumerate(models.Students.objects.all()[:50]): 
        list = [index + 1, student.roll_number, student.name, student.course, student.category,
                student.department,student.tuition_fee,student.insurance_fee,student.examination_fee,
                student.registration_fee,student.gymkhana_fee,student.medical_fee,student.student_benevolent_fund,
                student.lab_fee,student.semester_mess_advance,student.one_time_fee,student.refundable_security_deposit,
                student.accommodation_charges,student.student_welfare_fund,student.total_fee,student.mess_rebate,
                student.fee_payable - student.fee_arrear,student.fee_arrear,student.fee_payable]
        try: 
            payment = models.CurrentSemesterPayment.objects.get(student = student) 
            components = payment.components.all() 
            try: 
                comp = components.objects.get(mode = "payu") 
                list.extend([comp.mode,comp.type,"NA",comp.utr,comp.amt])
            except: 
                list.extend(["NA","NA","NA","NA",0]) 
            try: 
                comp = components.objects.get(mode = "neft") 
                list.extend([comp.mode,comp.type,"NA",comp.utr,comp.amt])
            except: 
                list.extend(["NA","NA","NA","NA",0]) 
            try: 
                comp = components.objects.get(mode = "upi") 
                list.extend([comp.mode,comp.type,"NA",comp.utr,comp.amt])
            except: 
                list.extend(["NA","NA","NA","NA",0]) 
            try: 
                comp = components.objects.get(mode = "qfix") 
                list.extend([comp.mode,comp.type,"NA",comp.utr,comp.amt])
            except: 
                list.extend(["NA","NA","NA","NA",0]) 
            print(list)
        except: 
            list.extend(["NA","NA","NA","NA",0])  
            list.extend(["NA","NA","NA","NA",0])  
            list.extend(["NA","NA","NA","NA",0])  
            list.extend(["NA","NA","NA","NA",0])  
        for col_index, value in enumerate(list, start=1):
            ws.cell(row=index + 3, column=col_index, value=value)


    print(ws.max_row)
    save_path = os.path.join(settings.BASE_DIR, "static", "excel", "download.xlsx")
    wb.save(save_path)
    return save_path



def loan_excel(excel_file): 
    col_range = "A:F" # sno,roll number,amt,mode,type,utr
    df = pd.read_excel(excel_file, usecols=col_range)
    cols = df.columns
    for _, row in df.iterrows():
        # try: 
        roll_number = row[cols[1]]
        student_instance = Students.objects.get(roll_number = roll_number)
        amt = int(row[cols[2]]) 
        mode = row[cols[3]]
        type = row[cols[4]] 
        utr = row[cols[5]] 
        student_instance.make_payment(amt,mode,type,utr)
        # except Exception as e: 
            # print(e)

def export_students():
    queryset = models.Students.objects.all()
    student_data = {
        "S. No.": list(range(1, len(queryset)+1)),
        "Roll Number": [student.roll_number for student in queryset],
        "Name": [student.name for student in queryset],
        "Course": [student.course for student in queryset],
        "Category": [student.category for student in queryset],
        "Department": [student.department for student in queryset],
        "Tuition Fee": [student.tuition_fee for student in queryset],
        "Insurance Fee": [student.insurance_fee for student in queryset],
        "Examination Fee": [student.examination_fee for student in queryset],
        "Registration Fee": [student.registration_fee for student in queryset],
        "Gymkhana Fee": [student.gymkhana_fee for student in queryset],
        "Medical Fee": [student.medical_fee for student in queryset],
        "Student Benevolent Fund": [student.student_benevolent_fund for student in queryset],
        "Lab Fee": [student.lab_fee for student in queryset],
        "Semester Mess Advance":[student.semester_mess_advance for student in queryset],
        "One Time Fee": [student.one_time_fee for student in queryset],
        "Refundable Security Deposit": [student.refundable_security_deposit for student in queryset],
        "Accommodation Charges": [student.accommodation_charges for student in queryset],
        "Student Welfare Fund": [student.student_welfare_fund for student in queryset],
        "Mess Rebate": [student.mess_rebate for student in queryset],
        "Fee Arrear": [student.fee_arrear for student in queryset],
        "Fee Payable": [student.fee_payable for student in queryset],
    }
    df = pd.DataFrame(student_data)
    return df 


def log(action: str):
    log_entry = models.Log(action=action)
    log_entry.save()


def set_remission(roll_number, remission_percentage):
    student = models.Students.objects.get(roll_number=roll_number)
    queryset = models.FeeRemission.objects.filter(student=student)
    if queryset.exists():
        fee_remission_instance = queryset[0]
        fee_remission_instance.percentage = remission_percentage
        fee_remission_instance.save()
    else:
        fee_remission_instance = models.FeeRemission(
            student=student, percentage=remission_percentage
        )
        fee_remission_instance.save()


def delete_remission(remission_instance: models.FeeRemission):
    student = remission_instance.student
    remission_instance.delete()


def excel_remission(excel_file):
    col_range = "A:C"
    df = pd.read_excel(excel_file, usecols=col_range)
    cols = df.columns
    for _, row in df.iterrows():
        set_remission(row[cols[1]], row[cols[2]])


def excel_delete(excel_file):
    df = pd.read_excel(excel_file)
    cols = df.columns
    success = 0
    fail = 0
    for _, row in df.iterrows():
        roll_number = row[cols[1]]
        try:
            student = models.Students.objects.get(roll_number=roll_number)
            student.delete()
            success += 1
        except:
            fail += 1
    return {'success' : success ,
            'fail' : fail}


# calculate fee structure of newly added student
def calculate_fee_structure(student: models.Students):
    course = student.course.split("-")[0]
    category = student.category
    fee_structure = models.FeeStructure.objects.get(
        course__icontains=course, category=category
    )
    assign_fee(student, fee_structure)


# recalculate fee structure when changed
def recalculate_fee_structure(fee_structure: models.FeeStructure):
    course = fee_structure.course
    category = fee_structure.category
    students = models.Students.objects.filter(
        course__icontains=course, category=category
    )
    for student in students:
        assign_fee(student, fee_structure)
    return len(students)


def assign_fee(student: models.Students, fee_structure: models.FeeStructure):
    student.base_tuition_fee = fee_structure.base_tuition_fee
    student.insurance_fee = fee_structure.insurance_fee
    student.examination_fee = fee_structure.examination_fee
    student.registration_fee = fee_structure.registration_fee
    student.gymkhana_fee = fee_structure.gymkhana_fee
    student.medical_fee = fee_structure.medical_fee
    student.student_benevolent_fund = fee_structure.student_benevolent_fund
    student.lab_fee = fee_structure.lab_fee
    student.semester_mess_advance = fee_structure.semester_mess_advance
    student.one_time_fee = fee_structure.one_time_fee
    student.refundable_security_deposit = fee_structure.refundable_security_deposit
    student.accommodation_charges = fee_structure.accommodation_charges
    student.student_welfare_fund = fee_structure.student_welfare_fund
    student.save()


def add_students(excel_file):
    col_range = "A:U"
    try:
        df = pd.read_excel(excel_file, usecols=col_range, skiprows=1, header=None)
        numerical_cols = list(range(6, 21))
        df[numerical_cols] = (
            df[numerical_cols]
            .apply(pd.to_numeric, errors="coerce")
            .fillna(0)
            .astype("int64")
        )
        print(df)
        for index, excel_row in df.iterrows():
            add_students_excel(excel_row)
    except Exception as e:
        print(e)
        return {"status": "error", "exception": e}

def add_students_excel(student_data):
    try:
        increment = 1
        student = models.Students(
            roll_number=student_data[0 + increment],
            name=student_data[1 + increment],
            course=student_data[2 + increment],
            category=student_data[3 + increment],
            department=student_data[4 + increment],
            base_tuition_fee=student_data[5 + increment],
            insurance_fee=student_data[6 + increment],
            examination_fee=student_data[7 + increment],
            registration_fee=student_data[8 + increment],
            gymkhana_fee=student_data[9 + increment],
            medical_fee=student_data[10 + increment],
            student_benevolent_fund=student_data[11 + increment],
            lab_fee=student_data[12 + increment],
            semester_mess_advance=student_data[13 + increment],
            one_time_fee=student_data[14 + increment],
            refundable_security_deposit=student_data[15 + increment],
            accommodation_charges=student_data[16 + increment],
            student_welfare_fund=student_data[17 + increment],
            mess_rebate=student_data[18 + increment],
            fee_arrear=student_data[19 + increment],
        )
        student.save()
    except Exception as e:
        print(e)

def add_students2(excel_file):
    col_range = "A:F"
    try:
        df = pd.read_excel(excel_file, usecols=col_range, skiprows=1, header=None)
        pd.set_option("display.max_columns", None)
        pd.set_option("display.max_rows", None)
        for index, excel_row in df.iterrows():
            add_students_excel2(excel_row)
    except Exception as e:
        print(e)
        return {"status": "error", "exception": e}

def add_students_excel2(student_data):
    try:
        increment = 1
        student = models.Students(
            roll_number=student_data[0 + increment],
            name=student_data[1 + increment],
            course=student_data[2 + increment],
            category=student_data[3 + increment],
            department=student_data[4 + increment],
        )
        calculate_fee_structure(student)
        student.save()
    except Exception as e:
        print(e)
